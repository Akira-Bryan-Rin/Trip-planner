from llm import duckchat_response
from datetime import datetime, timedelta
from IPython.display import display, HTML
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# 模擬 location_description 函式
def location_description(name):
    prompt = f"請輸出一篇大約五十字的{name}的景點介紹，不用前言只需本文"
    brief_intro = duckchat_response(prompt)       
    
    return brief_intro

def schedule_brief(locations):
    prompt = f"請輸出一篇大約兩百字的行程簡介將各景點依造訪時間串聯，{locations}為行程中需造訪的景點。"

    try:
        duckchat_response(prompt)[duckchat_response(prompt).find(":")+1:].strip() 
    except Exception as e:
        brief_schedule = duckchat_response(prompt)
    else:
        brief_schedule = duckchat_response(prompt)[duckchat_response(prompt).find(":")+1:].strip() 

    return brief_schedule

def calculate_stay_times(travel_times):
    """
    根據地點間的移動時間計算地點1~6的逗留時間，並以字串形式返回。

    Args:
        travel_times (list): 長度為5的列表，表示地點1到地點6間的移動時間（分鐘）。

    Returns:
        list: 地點1~6的逗留時間段字串，格式為 "09:00~11:30"。
    """
    if len(travel_times) != 5:
        raise ValueError("移動時間列表必須包含5個元素，表示地點1到地點6間的移動時間。")

    # 初始條件
    start_location_1 = datetime.strptime("09:00", "%H:%M")
    arrive_location_2_fixed = datetime.strptime("12:00", "%H:%M")
    leave_location_2 = arrive_location_2_fixed + timedelta(hours=1)  # 地點2固定逗留1小時
    arrive_location_5_fixed = datetime.strptime("18:00", "%H:%M")  # 地點5的到達時間固定為18:00
    stay_duration_location_5 = timedelta(hours=1)  # 地點5固定逗留1小時

    # 計算地點3與地點4的中點時間
    midpoint_location_3_4 = leave_location_2 + (arrive_location_5_fixed - leave_location_2) / 2

    # 地點1的離開時間
    leave_location_1 = arrive_location_2_fixed - timedelta(minutes=travel_times[0])
    stay_times = [f"{start_location_1.strftime('%H:%M')}~{leave_location_1.strftime('%H:%M')}"]

    # 地點2的逗留時間
    stay_times.append(f"{arrive_location_2_fixed.strftime('%H:%M')}~{leave_location_2.strftime('%H:%M')}")

    # 地點3的逗留時間
    arrive_location_3 = leave_location_2 + timedelta(minutes=travel_times[1])
    leave_location_3 = midpoint_location_3_4 - timedelta(minutes=travel_times[2] // 2)
    stay_times.append(f"{arrive_location_3.strftime('%H:%M')}~{leave_location_3.strftime('%H:%M')}")

    # 地點4的逗留時間
    arrive_location_4 = midpoint_location_3_4 + timedelta(minutes=travel_times[2] // 2)
    leave_location_4 = arrive_location_5_fixed - timedelta(minutes=travel_times[3])
    stay_times.append(f"{arrive_location_4.strftime('%H:%M')}~{leave_location_4.strftime('%H:%M')}")

    # 地點5的逗留時間
    stay_times.append(f"{arrive_location_5_fixed.strftime('%H:%M')}~{(arrive_location_5_fixed + stay_duration_location_5).strftime('%H:%M')}")

    # 地點6的到達時間
    arrive_location_6 = arrive_location_5_fixed + stay_duration_location_5 + timedelta(minutes=travel_times[4])
    stay_times.append(f"{arrive_location_6.strftime('%H:%M')}~")

    return stay_times

# 處理換行標記：/n
# 使用 HTML 格式化 DataFrame
def display_dataframe_with_linebreaks(df):
    # 將 DataFrame 轉為 HTML 並嵌入樣式
    styled_table = df.to_html(
        index=False, escape=False  # 保持 HTML 標籤（如 <br>）有效
    ).replace(
        "<td>", '<td style="text-align: left; white-space: pre-wrap; word-break: break-word;">'  # 為每個 <td> 添加樣式
    ).replace(
        "<th>", '<th style="text-align: right;">'  # 為每個 <th> 添加樣式，靠右對齊
    )
    display(HTML(styled_table))
    
def set_column_widths(excel_file, column_widths):
    """
    調整 Excel 檔案的欄位寬度。
    
    Args:
        excel_file (str): Excel 檔案路徑。
        column_widths (dict): 指定每欄的寬度，格式為 {列名: 寬度}。
    """
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    
    for col_idx, width in column_widths.items():
        column_letter = get_column_letter(col_idx)  # 轉換列索引為字母（例如 1 -> A）
        sheet.column_dimensions[column_letter].width = width
    
    workbook.save(excel_file)
    
def create_travel_schedule(locations,travel_times):
    """
    接收6個地點名稱，生成行程表，輸出至 Excel 並在終端顯示。
    
    Args:
        locations (list): 包含6個地點名稱的列表。
    """
    if len(locations) != 6:
        raise ValueError("請確保輸入的地點數量正好是 6 個。")

    # 初始化一個空的列表來存放地點資訊
    schedule = []
    schedulebrief = schedule_brief(locations)

    for i, name in enumerate(locations):
        # 使用 location_description 為所有地點生成簡介存入description
        description = location_description(name)
        # 把travel times計算後存入time_period
        time_periods = calculate_stay_times(travel_times)
        
        
        # 將地點資訊加入行程表
        schedule.append({"時間": time_periods[i], "地點名稱": f"{name[0]}：{name[3]}", "景點簡介": description})

    # 使用 Pandas 創建 DataFrame，並enable換行顯示
    pd.set_option("expand_frame_repr", True)
    pd.set_option('max_colwidth',500)

    df = pd.DataFrame(schedule)
    
    # 將行程摘要輸出到行程表下方
    sepr_row = pd.DataFrame({'時間': ['----'], '地點名稱': [''], '景點簡介': ['']}, index=[0])
    schd_row = pd.DataFrame({'時間': ['行程摘要'], '地點名稱': [''], '景點簡介': [schedulebrief]}, index=[0]) 
    
    df = pd.concat([df, sepr_row, schd_row], ignore_index=True)
    
    # 如果數據中原本是用 \n 表示換行，先替換為 <br>
    df = df.replace("\n", "<br>", regex=True)
    # 再次格式化顯示
    styled_df = df.style.set_table_styles([{'selector': 'th', 'props': [('text-align', 'left')]},  {'selector': 'td', 'props': [('text-align', 'left')]}])
    styled_df = styled_df.hide(axis='index')
    display_dataframe_with_linebreaks(styled_df)

    """
    新增旅遊行程表到 Excel 文件，資料貼在現有內容的下方，並間隔兩行。

    :param df: 新的 DataFrame，包含要新增的資料
    """
    # 定義檔案路徑，將檔案存放於桌面
    excel_file = r"travel_schedule.xlsx"
    sheet_name = "行程表"

    # 檢查 Excel 文件是否存在
    if os.path.exists(excel_file):
        # 如果文件存在，讀取現有內容
        existing_data = pd.read_excel(excel_file, sheet_name=sheet_name)

        # 找到目前的「第幾天」
        last_day_row = existing_data.loc[existing_data.iloc[:, 0].str.contains("第", na=False, regex=True)].tail(1)
        if not last_day_row.empty:
            # 取出最後的天數，假設格式為「第1天」
            last_day = int(last_day_row.iloc[0, 0][1:-1])  # 提取數字部分
        else:
            last_day = 0  # 如果找不到標記，從第0天開始
    else:
        # 如果文件不存在，從第1天開始
        existing_data = pd.DataFrame()  # 空的 DataFrame
        last_day = 0

    # 設定新資料的標記
    new_day = last_day + 1
    day_marker = pd.DataFrame([[f"第{new_day}天"] + [""] * (df.shape[1] - 1)], columns=df.columns)

    # 合併新資料與標記
    if not existing_data.empty:
        spacer_rows = pd.DataFrame([[""] * existing_data.shape[1]] * 2, columns=existing_data.columns)  # 兩行間隔
        combined_data = pd.concat([existing_data, spacer_rows, day_marker, df], ignore_index=True)
    else:
        combined_data = pd.concat([day_marker, df], ignore_index=True)  # 如果沒有舊資料，直接新增

    # 將結果寫回到 Excel
    with pd.ExcelWriter(excel_file, engine="openpyxl", mode="w") as writer:
        combined_data.to_excel(writer, index=False, sheet_name=sheet_name)
    
    # print(f"\n行程表已輸出至 {excel_file}\n")

    # 調整EExcel輸出欄位寬度
    column_widths = {
        1: 15,  # "時段"列
        2: 20,  # "地點名稱"列
        3: 50   # "景點簡介"列
    }
    set_column_widths(excel_file, column_widths)